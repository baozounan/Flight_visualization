import pandas as pd
import csv
import openpyxl
import numpy as np
import xlsxwriter as xw
def str2dict(str):

    """将str类型转换成字典类型，输入一个字符串，返回一个字典"""

    dict = exec(str)
    return

def savedata2xlsx(data):
    """将data保存为本地xlsx文件"""
    new = openpyxl.load_workbook("E:\\test_data\\test.xlsx")  # 加载文件
    sheet = new.create_sheet(title="sheet_title")
    for row in range(len(data)):  # len(use_data)
       # len(use_title)
         _ = sheet.cell(row=row + 1, column=1, value=u'%s' % data[0][row])
    newexcel = new.save("E:\\test_data\\test.xlsx")
    return None


def xw_toexcel(title,data,filename): # xlsxwriter库储存数据到excel
    """写入数据到本地，filename=文件位置，data是写入的数据是多维数组，title是表格中的列名"""
    data = np.array(data)
    workbook = xw.Workbook(filename) # 创建工作簿
    worksheet1 = workbook.add_worksheet("sheet1") # 创建子表
    worksheet1.activate() # 激活表
    worksheet1.write_row('A1',title) # 从A1单元格开始写入表头
    i = 2 # 从第二行开始写入数据
    for j in range(len(data)):
        insertData = data[:,j]  # 取一列数据
        row = 'A' + str(i)
        worksheet1.write_row(row, insertData)
        i += 1
    workbook.close() # 关闭表
    return None

def appendCsvData(data):

    """将数据追加到csv文件中"""
    file=open("data\\panel_test_data.csv","a+",newline="")
    reader = csv.reader(file)
    original = str(reader)
    content = csv.writer(file)
    for row in original:
       content.writerow(row)
    content.writerow(data)
    file.close()
    return None






